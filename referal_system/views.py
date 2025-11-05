from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib.auth import login, authenticate, logout
from django.contrib import messages
from django.contrib.auth.models import User
from django.db.models import Sum, Count
from .models import Partner, Lead, LeadStage, Transaction, Payout
from django.utils import timezone
from django.db.models import Sum, Count, Q

# Partner Login (FIXED - Proper Team Member Check)
def partner_login(request):
    if request.user.is_authenticated:
        if request.user.is_staff:
            return redirect('admin_dashboard')
        
        # Check in order: Partner → Team Member
        try:
            if hasattr(request.user, 'partner') and request.user.partner:
                return redirect('partner_dashboard')
        except:
            pass
        
        try:
            if hasattr(request.user, 'teammember') and request.user.teammember:
                return redirect('team_dashboard')
        except:
            pass
    
    if request.method == 'POST':
        username = request.POST.get('username')
        password = request.POST.get('password')
        user = authenticate(request, username=username, password=password)
        
        if user is not None:
            login(request, user)  # Login first
            
            # Then redirect based on user type
            if user.is_staff:
                return redirect('admin_dashboard')
            
            # Check Partner
            try:
                if hasattr(user, 'partner') and user.partner:
                    return redirect('partner_dashboard')
            except:
                pass
            
            # Check Team Member
            try:
                if hasattr(user, 'teammember') and user.teammember:
                    return redirect('team_dashboard')
            except:
                pass
            
            # If no role found
            logout(request)
            messages.error(request, 'You are not registered as Partner or Team Member')
        else:
            messages.error(request, 'Invalid username or password')
    
    return render(request, 'referal_system/partner_login.html')


# Partner Logout
def partner_logout(request):
    logout(request)
    return redirect('partner_login')

# Partner Dashboard
@login_required
def partner_dashboard(request):
    if request.user.is_staff:
        return redirect('admin_dashboard')
    
    partner = get_object_or_404(Partner, user=request.user)
    
    total_leads = Lead.objects.filter(partner=partner).count()
    total_commission = Lead.objects.filter(partner=partner, commission_paid=True).aggregate(
        total=Sum('commission_amount'))['total'] or 0
    pending_commission = Lead.objects.filter(partner=partner, commission_paid=False).aggregate(
        total=Sum('commission_amount'))['total'] or 0
    
    recent_leads = Lead.objects.filter(partner=partner).order_by('-created_at')[:10]
    recent_transactions = Transaction.objects.filter(partner=partner).order_by('-created_at')[:10]
    
    context = {
        'partner': partner,
        'total_leads': total_leads,
        'total_commission': total_commission,
        'pending_commission': pending_commission,
        'recent_leads': recent_leads,
        'recent_transactions': recent_transactions,
    }
    
    return render(request, 'referal_system/partner_dashboard.html', context)

# Partner Add Lead
# Partner Add Lead (UPDATED - Own vs Referral)
# Partner Add Lead (UPDATED - created_by hataya)
@login_required
def partner_add_lead(request):
    if request.user.is_staff:
        return redirect('admin_dashboard')
    
    partner = get_object_or_404(Partner, user=request.user)
    stages = LeadStage.objects.all()
    
    if request.method == 'POST':
        customer_name = request.POST.get('customer_name')
        customer_email = request.POST.get('customer_email')
        customer_phone = request.POST.get('customer_phone')
        stage_id = request.POST.get('stage')
        deal_amount = request.POST.get('deal_amount', 0)
        notes_text = request.POST.get('notes', '')  # Rename kar diya
        
        lead_ownership = request.POST.get('lead_ownership')
        stage = LeadStage.objects.get(id=stage_id) if stage_id else None
        
        if lead_ownership == 'own':
            # Lead create karo WITHOUT notes
            lead = Lead.objects.create(
                lead_type='partner_own',
                partner=partner,
                customer_name=customer_name,
                customer_email=customer_email,
                customer_phone=customer_phone,
                stage=stage,
                deal_amount=float(deal_amount) if deal_amount else 0,
                assigned_to_admin=False
            )
            
            # Agar notes text hai to LeadNote create karo
            if notes_text:
                LeadNote.objects.create(
                    lead=lead,
                    note=notes_text,
                    created_by=request.user
                )
            
            messages.success(request, 'Own lead added successfully! You can manage this lead.')
        else:
            # Referral lead
            lead = Lead.objects.create(
                lead_type='partner_referral',
                partner=partner,
                customer_name=customer_name,
                customer_email=customer_email,
                customer_phone=customer_phone,
                stage=stage,
                deal_amount=float(deal_amount) if deal_amount else 0,
                assigned_to_admin=True
            )
            
            # Notes add karo agar hai
            if notes_text:
                LeadNote.objects.create(
                    lead=lead,
                    note=notes_text,
                    created_by=request.user
                )
            
            messages.success(request, 'Referral lead submitted to admin! You will receive commission when closed.')
        
        return redirect('partner_leads')
    
    context = {
        'partner': partner,
        'stages': stages,
    }
    
    return render(request, 'referal_system/partner_add_lead.html', context)


# Partner Leads List
# Partner Leads List (UPDATED)
# Partner Leads List
@login_required
def partner_leads(request):
    if request.user.is_staff:
        return redirect('admin_dashboard')
    
    partner = get_object_or_404(Partner, user=request.user)
    leads = Lead.objects.filter(partner=partner).order_by('-created_at')
    stages = LeadStage.objects.all()  # YE LINE ADD HUI
    
    context = {
        'partner': partner,
        'leads': leads,
        'stages': stages,  # YE LINE ADD HUI
    }
    
    return render(request, 'referal_system/partner_leads.html', context)



# Partner Wallet
@login_required
def partner_wallet(request):
    if request.user.is_staff:
        return redirect('admin_dashboard')
    
    partner = get_object_or_404(Partner, user=request.user)
    transactions = Transaction.objects.filter(partner=partner).order_by('-created_at')
    payouts = Payout.objects.filter(partner=partner).order_by('-requested_at')
    
    context = {
        'partner': partner,
        'transactions': transactions,
        'payouts': payouts,
    }
    
    return render(request, 'referal_system/partner_wallet.html', context)

# Partner Request Payout
@login_required
def partner_request_payout(request):
    if request.user.is_staff:
        return redirect('admin_dashboard')
    
    partner = get_object_or_404(Partner, user=request.user)
    
    if request.method == 'POST':
        amount = float(request.POST.get('amount'))
        bank_details = request.POST.get('bank_details')
        
        if amount <= 0:
            messages.error(request, 'Please enter a valid amount!')
            return redirect('partner_wallet')
        
        if amount > partner.wallet_balance:
            messages.error(request, 'Insufficient balance!')
            return redirect('partner_wallet')
        
        Payout.objects.create(
            partner=partner,
            amount=amount,
            bank_details=bank_details,
            status='pending'
        )
        
        messages.success(request, 'Payout request submitted successfully!')
        return redirect('partner_wallet')
    
    context = {
        'partner': partner,
    }
    
    return render(request, 'referal_system/partner_request_payout.html', context)

# ===========================================
# ADMIN VIEWS
# ===========================================

# Admin Dashboard
@login_required
def admin_dashboard(request):
    if not request.user.is_staff:
        messages.error(request, 'You do not have permission to access admin dashboard')
        return redirect('partner_login')
    
    total_partners = Partner.objects.count()
    total_leads = Lead.objects.count()
    admin_leads = Lead.objects.filter(lead_type='admin').count()
    partner_leads = Lead.objects.filter(lead_type='partner').count()
    
    total_commission = Lead.objects.filter(commission_paid=True).aggregate(
        total=Sum('commission_amount'))['total'] or 0
    pending_payouts = Payout.objects.filter(status='pending').count()
    
    recent_leads = Lead.objects.all().order_by('-created_at')[:10]
    recent_payouts = Payout.objects.all().order_by('-requested_at')[:5]
    
    context = {
        'total_partners': total_partners,
        'total_leads': total_leads,
        'admin_leads': admin_leads,
        'partner_leads': partner_leads,
        'total_commission': total_commission,
        'pending_payouts': pending_payouts,
        'recent_leads': recent_leads,
        'recent_payouts': recent_payouts,
    }
    
    return render(request, 'referal_system/admin_dashboard.html', context)

from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.contrib.auth.models import User
from django.urls import reverse
from .models import Partner

# Admin Partners List
@login_required
def admin_partners(request):
    if not request.user.is_staff:
        return redirect('partner_login')
    
    if request.method == 'POST':
        # Register new partner
        name = request.POST.get('name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        # Additional fields
        company_name = request.POST.get('company_name', '')
        address = request.POST.get('address', '')
        city = request.POST.get('city', '')
        state = request.POST.get('state', '')
        pincode = request.POST.get('pincode', '')
        
        # KYC Details
        aadhaar_number = request.POST.get('aadhaar_number', '')
        pan_number = request.POST.get('pan_number', '')
        gst_number = request.POST.get('gst_number', '')
        
        # Bank Details
        bank_name = request.POST.get('bank_name', '')
        account_number = request.POST.get('account_number', '')
        ifsc_code = request.POST.get('ifsc_code', '')
        account_holder_name = request.POST.get('account_holder_name', '')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists!')
        elif Partner.objects.filter(email=email).exists():
            messages.error(request, 'Email already exists!')
        else:
            user = User.objects.create_user(username=username, password=password, email=email)
            Partner.objects.create(
                user=user, 
                name=name, 
                email=email, 
                phone=phone,
                company_name=company_name,
                address=address,
                city=city,
                state=state,
                pincode=pincode,
                aadhaar_number=aadhaar_number,
                pan_number=pan_number,
                gst_number=gst_number,
                bank_name=bank_name,
                account_number=account_number,
                ifsc_code=ifsc_code,
                account_holder_name=account_holder_name
            )
            messages.success(request, 'Partner registered successfully!')
            return redirect('admin_partners')
    
    partners = Partner.objects.all().order_by('-created_at')
    
    # Generate registration link
    registration_url = request.build_absolute_uri(reverse('partner_register'))
    
    context = {
        'partners': partners,
        'registration_url': registration_url,
    }
    
    return render(request, 'referal_system/admin_partners.html', context)


# Edit Partner
@login_required
def admin_edit_partner(request, partner_id):
    if not request.user.is_staff:
        return redirect('partner_login')
    
    partner = get_object_or_404(Partner, id=partner_id)
    
    if request.method == 'POST':
        # Update partner details
        partner.name = request.POST.get('name')
        partner.email = request.POST.get('email')
        partner.phone = request.POST.get('phone')
        
        # Additional fields
        partner.company_name = request.POST.get('company_name', '')
        partner.address = request.POST.get('address', '')
        partner.city = request.POST.get('city', '')
        partner.state = request.POST.get('state', '')
        partner.pincode = request.POST.get('pincode', '')
        
        # KYC Details
        partner.aadhaar_number = request.POST.get('aadhaar_number', '')
        partner.pan_number = request.POST.get('pan_number', '')
        partner.gst_number = request.POST.get('gst_number', '')
        
        # Bank Details
        partner.bank_name = request.POST.get('bank_name', '')
        partner.account_number = request.POST.get('account_number', '')
        partner.ifsc_code = request.POST.get('ifsc_code', '')
        partner.account_holder_name = request.POST.get('account_holder_name', '')
        
        # Status
        partner.is_active = request.POST.get('is_active') == 'on'
        partner.is_verified = request.POST.get('is_verified') == 'on'
        
        # Update username if changed
        new_username = request.POST.get('username')
        if new_username != partner.user.username:
            if User.objects.filter(username=new_username).exists():
                messages.error(request, 'Username already exists!')
                return redirect('admin_edit_partner', partner_id=partner_id)
            partner.user.username = new_username
            partner.user.save()
        
        # Update password if provided
        new_password = request.POST.get('password')
        if new_password:
            partner.user.set_password(new_password)
            partner.user.save()
        
        partner.save()
        messages.success(request, 'Partner updated successfully!')
        return redirect('admin_partners')
    
    context = {
        'partner': partner,
    }
    
    return render(request, 'referal_system/admin_edit_partner.html', context)


# Delete Partner
@login_required
def admin_delete_partner(request, partner_id):
    if not request.user.is_staff:
        return redirect('partner_login')
    
    partner = get_object_or_404(Partner, id=partner_id)
    
    if request.method == 'POST':
        partner_name = partner.name
        user = partner.user
        partner.delete()
        user.delete()
        messages.success(request, f'Partner "{partner_name}" deleted successfully!')
        return redirect('admin_partners')
    
    context = {
        'partner': partner,
    }
    
    return render(request, 'referal_system/admin_delete_partner.html', context)


# Toggle Partner Status (Quick Active/Inactive)
@login_required
def admin_toggle_partner_status(request, partner_id):
    if not request.user.is_staff:
        return redirect('partner_login')
    
    partner = get_object_or_404(Partner, id=partner_id)
    partner.is_active = not partner.is_active
    partner.save()
    
    status = "activated" if partner.is_active else "deactivated"
    messages.success(request, f'Partner "{partner.name}" {status} successfully!')
    return redirect('admin_partners')


# Public Partner Registration
def partner_register(request):
    if request.method == 'POST':
        name = request.POST.get('name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        username = request.POST.get('username')
        password = request.POST.get('password')
        
        # Additional fields
        company_name = request.POST.get('company_name', '')
        address = request.POST.get('address', '')
        city = request.POST.get('city', '')
        state = request.POST.get('state', '')
        pincode = request.POST.get('pincode', '')
        
        # KYC Details
        aadhaar_number = request.POST.get('aadhaar_number', '')
        pan_number = request.POST.get('pan_number', '')
        gst_number = request.POST.get('gst_number', '')
        
        # Bank Details
        bank_name = request.POST.get('bank_name', '')
        account_number = request.POST.get('account_number', '')
        ifsc_code = request.POST.get('ifsc_code', '')
        account_holder_name = request.POST.get('account_holder_name', '')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists!')
        elif Partner.objects.filter(email=email).exists():
            messages.error(request, 'Email already exists!')
        else:
            user = User.objects.create_user(username=username, password=password, email=email)
            Partner.objects.create(
                user=user, 
                name=name, 
                email=email, 
                phone=phone,
                company_name=company_name,
                address=address,
                city=city,
                state=state,
                pincode=pincode,
                aadhaar_number=aadhaar_number,
                pan_number=pan_number,
                gst_number=gst_number,
                bank_name=bank_name,
                account_number=account_number,
                ifsc_code=ifsc_code,
                account_holder_name=account_holder_name
            )
            messages.success(request, 'Registration successful! Please wait for admin verification.')
            return redirect('partner_login')
    
    return render(request, 'referal_system/partner_register.html')


# Admin All Leads
# Admin All Leads (UPDATED - Sirf assigned leads dikhengi)
# Admin All Leads (UPDATED - Sirf assigned leads dikhengi)
# Update admin_leads view
from datetime import datetime, timedelta
from django.utils import timezone

@login_required
def admin_leads(request):
    if not request.user.is_staff:
        return redirect('partner_login')
    
    # Base queryset
    leads = Lead.objects.filter(
        Q(lead_type='admin') |
        Q(lead_type='partner_referral') |
        Q(lead_type='partner_own', assigned_to_admin=True)
    )
    
    # ✅ FILTER: Lead Type
    lead_type_filter = request.GET.get('lead_type')
    if lead_type_filter:
        leads = leads.filter(lead_type=lead_type_filter)
    
    # ✅ FILTER: Partner
    partner_filter = request.GET.get('partner')
    if partner_filter:
        leads = leads.filter(partner_id=partner_filter)
    
    # ✅ FILTER: Follow-up Date
    follow_up_filter = request.GET.get('follow_up_filter')
    today = timezone.now().date()
    
    if follow_up_filter == 'today':
        leads = leads.filter(next_follow_up=today)
    elif follow_up_filter == 'tomorrow':
        tomorrow = today + timedelta(days=1)
        leads = leads.filter(next_follow_up=tomorrow)
    elif follow_up_filter == 'this_week':
        week_end = today + timedelta(days=7)
        leads = leads.filter(next_follow_up__range=[today, week_end])
    elif follow_up_filter == 'overdue':
        leads = leads.filter(next_follow_up__lt=today)
    
    leads = leads.order_by('-created_at')
    
    stages = LeadStage.objects.all()
    partners = Partner.objects.filter(is_active=True)
    team_members = TeamMember.objects.filter(status='active')
    
    context = {
        'leads': leads,
        'stages': stages,
        'partners': partners,
        'team_members': team_members,
    }
    
    return render(request, 'referal_system/admin_leads.html', context)

# Admin Update Lead
# Admin Update Lead (Admin sirf referral leads aur admin leads update kar sakta hai)
# Admin Update Lead (UPDATED - Partner own assigned leads update kar sakta hai)
from decimal import Decimal

from decimal import Decimal

from decimal import Decimal
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required

from decimal import Decimal
from django.contrib import messages
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required

@login_required
def admin_add_lead(request):
    if not request.user.is_staff:
        return redirect('partner_login')
    
    stages = LeadStage.objects.all()
    partners = Partner.objects.filter(is_active=True)
    
    if request.method == 'POST':
        lead_type_raw = request.POST.get('lead_type')  # 'admin' or 'partner' from form
        partner_id = request.POST.get('partner')
        customer_name = request.POST.get('customer_name')
        customer_email = request.POST.get('customer_email')
        customer_phone = request.POST.get('customer_phone')
        stage_id = request.POST.get('stage')
        deal_amount = request.POST.get('deal_amount', 0)
        commission_percent = request.POST.get('commission_percent', 0)
        
        # 🔍 DEBUGGING
        print("=" * 50)
        print("🔍 DEBUG - Admin Add Lead")
        print(f"Lead Type (raw from form): {lead_type_raw}")
        print(f"Partner ID: {partner_id}")
        print(f"Deal Amount (raw): {deal_amount}")
        print(f"Commission Percent (raw): {commission_percent}")
        
        # ✅ FIX: Map form lead_type to model lead_type
        if lead_type_raw == 'admin':
            lead_type = 'admin'
            partner = None
        else:  # lead_type_raw == 'partner'
            lead_type = 'partner_referral'  # ✅ YAHAN CHANGE KIYA
            partner = Partner.objects.get(id=partner_id) if partner_id else None
        
        print(f"Lead Type (mapped for model): {lead_type}")
        
        stage = LeadStage.objects.get(id=stage_id) if stage_id else None
        
        # ✅ Convert to Decimal
        deal_amount_decimal = Decimal(deal_amount) if deal_amount else Decimal('0')
        commission_percent_decimal = Decimal(commission_percent) if commission_percent else Decimal('0')
        
        print(f"Deal Amount (Decimal): {deal_amount_decimal}")
        print(f"Commission Percent (Decimal): {commission_percent_decimal}")
        
        # ✅ Create lead
        new_lead = Lead.objects.create(
            lead_type=lead_type,
            partner=partner,
            customer_name=customer_name,
            customer_email=customer_email,
            customer_phone=customer_phone,
            stage=stage,
            deal_amount=deal_amount_decimal,
            commission_percent=commission_percent_decimal,
            assigned_to_admin=True
        )
        
        # 🔍 DEBUGGING: Check saved values
        print(f"✅ Lead Created - ID: {new_lead.id}")
        print(f"Commission Amount (calculated): {new_lead.commission_amount}")
        print("=" * 50)
        
        messages.success(request, f'✅ Lead added successfully! Commission: ₹{new_lead.commission_amount}')
        return redirect('admin_leads')
    
    context = {
        'stages': stages,
        'partners': partners,
    }
    
    return render(request, 'referal_system/admin_add_lead.html', context)


@login_required
def admin_update_lead(request, lead_id):
    if not request.user.is_staff:
        return redirect('partner_login')
    
    lead = get_object_or_404(Lead, id=lead_id)
    
    # ✅ CHECK: Partner own leads jo assigned nahi hain
    if lead.lead_type == 'partner_own' and not lead.assigned_to_admin:
        messages.error(request, '❌ You cannot update this lead. Partner has not assigned it yet!')
        return redirect('admin_leads')
    
    # ✅ GET REQUEST: Show update form
    if request.method == 'GET':
        stages = LeadStage.objects.all()
        partners = Partner.objects.filter(is_active=True)
        
        context = {
            'lead': lead,
            'stages': stages,
            'partners': partners,
        }
        return render(request, 'referal_system/admin_update_lead.html', context)
    
    # ✅ POST REQUEST: Update lead
    if request.method == 'POST':
        # 🔍 DEBUGGING: Print old values
        print("=" * 50)
        print("🔍 DEBUG - Admin Update Lead")
        print(f"Lead ID: {lead.id}")
        print(f"Old Customer Name: {lead.customer_name}")
        print(f"Old Deal Amount: {lead.deal_amount}")
        print(f"Old Commission Percent: {lead.commission_percent}")
        print(f"Old Commission Amount: {lead.commission_amount}")
        print(f"Old Commission Paid: {lead.commission_paid}")
        
        # ✅ Get form data
        customer_name = request.POST.get('customer_name', lead.customer_name)
        customer_email = request.POST.get('customer_email', lead.customer_email)
        customer_phone = request.POST.get('customer_phone', lead.customer_phone)
        stage_id = request.POST.get('stage')
        deal_amount = request.POST.get('deal_amount', 0)
        commission_percent = request.POST.get('commission_percent', 0)
        commission_paid = request.POST.get('commission_paid') == 'on'
        
        print(f"New Customer Name: {customer_name}")
        print(f"New Deal Amount (raw): {deal_amount}")
        print(f"New Commission Percent (raw): {commission_percent}")
        print(f"Commission Paid Checkbox: {commission_paid}")
        
        # ✅ Update customer details
        lead.customer_name = customer_name
        lead.customer_email = customer_email
        lead.customer_phone = customer_phone
        
        # ✅ Update stage
        if stage_id:
            lead.stage = LeadStage.objects.get(id=stage_id)
        
        # ✅ Update deal amount
        lead.deal_amount = Decimal(deal_amount) if deal_amount else Decimal('0')
        
        # ✅ Update commission (for partner_referral AND admin leads)
        if lead.lead_type in ['partner_referral', 'admin']:
            lead.commission_percent = Decimal(commission_percent) if commission_percent else Decimal('0')
            
            print(f"New Deal Amount (Decimal): {lead.deal_amount}")
            print(f"New Commission Percent (Decimal): {lead.commission_percent}")
        
        # ✅✅ FIRST SAVE TO CALCULATE COMMISSION AMOUNT
        lead.save()  # Yahan save() se commission_amount calculate hoga!
        
        print(f"After Save - Commission Amount: {lead.commission_amount}")
        
        # ✅ NOW handle commission payment (only for partner_referral)
        if lead.lead_type == 'partner_referral' and commission_paid and not lead.commission_paid and lead.partner:
            print(f"💰 Processing Commission Payment...")
            print(f"Commission Amount to Credit: ₹{lead.commission_amount}")
            print(f"Partner: {lead.partner.name}")
            print(f"Old Wallet Balance: ₹{lead.partner.wallet_balance}")
            
            # ✅ Mark commission as paid
            lead.commission_paid = True
            
            # ✅ Wallet credit
            lead.partner.wallet_balance += lead.commission_amount
            lead.partner.save()
            
            print(f"New Wallet Balance: ₹{lead.partner.wallet_balance}")
            
            # ✅ Transaction create
            Transaction.objects.create(
                partner=lead.partner,
                lead=lead,
                transaction_type='credit',
                amount=lead.commission_amount,
                description=f'Commission for referral lead: {lead.customer_name}'
            )
            
            # ✅ Save lead again to update commission_paid status
            lead.save()
            
            print(f"✅ Commission Paid: ₹{lead.commission_amount} to {lead.partner.name}")
            messages.success(request, f'💰 Commission of ₹{lead.commission_amount} credited to {lead.partner.name}\'s wallet!')
        
        # 🔍 DEBUGGING: Check final values
        print(f"✅ Lead Updated Successfully")
        print(f"Final Deal Amount: {lead.deal_amount}")
        print(f"Final Commission Percent: {lead.commission_percent}")
        print(f"Final Commission Amount: {lead.commission_amount}")
        print(f"Final Commission Paid: {lead.commission_paid}")
        print("=" * 50)
        
        messages.success(request, '✅ Lead updated successfully!')
        return redirect('admin_leads')
    
# Admin Lead Stages
# Admin Lead Stages (UPDATED - created_by hataya)
@login_required
def admin_stages(request):
    if not request.user.is_staff:
        return redirect('partner_login')
    
    if request.method == 'POST':
        name = request.POST.get('name')
        order = request.POST.get('order', 0)
        
        LeadStage.objects.create(
            name=name,
            order=int(order),
            created_by=request.user  # Agar created_by field hai model mein
        )
        
        messages.success(request, 'Stage created successfully!')
        return redirect('admin_stages')
    
    stages = LeadStage.objects.all().order_by('order')
    
    context = {
        'stages': stages,
    }
    
    return render(request, 'referal_system/admin_stages.html', context)


@login_required
def admin_edit_stage(request, stage_id):
    if not request.user.is_staff:
        return redirect('partner_login')
    
    stage = get_object_or_404(LeadStage, id=stage_id)
    
    if request.method == 'POST':
        stage.name = request.POST.get('name')
        stage.order = int(request.POST.get('order', 0))
        stage.save()
        
        messages.success(request, 'Stage updated successfully!')
        return redirect('admin_stages')
    
    return redirect('admin_stages')


@login_required
def admin_delete_stage(request, stage_id):
    if not request.user.is_staff:
        return redirect('partner_login')
    
    stage = get_object_or_404(LeadStage, id=stage_id)
    stage.delete()
    
    messages.success(request, 'Stage deleted successfully!')
    return redirect('admin_stages')
# Admin Payouts
@login_required
def admin_payouts(request):
    if not request.user.is_staff:
        return redirect('partner_login')
    
    payouts = Payout.objects.all().order_by('-requested_at')
    
    context = {
        'payouts': payouts,
    }
    
    return render(request, 'referal_system/admin_payouts.html', context)

# Admin Process Payout
@login_required
def admin_process_payout(request, payout_id):
    if not request.user.is_staff:
        return redirect('partner_login')
    
    payout = get_object_or_404(Payout, id=payout_id)
    
    if request.method == 'POST':
        action = request.POST.get('action')
        remarks = request.POST.get('remarks', '')
        
        if action == 'approve':
            payout.status = 'approved'
            payout.processed_by = request.user
            payout.processed_at = timezone.now()
            payout.remarks = remarks
            payout.save()
            messages.success(request, 'Payout approved!')
            
        elif action == 'complete':
            if payout.status == 'approved':
                payout.status = 'completed'
                payout.processed_by = request.user
                payout.processed_at = timezone.now()
                payout.remarks = remarks
                payout.save()
                
                # Wallet se deduct karo
                payout.partner.wallet_balance -= payout.amount
                payout.partner.save()
                
                # Transaction create karo
                Transaction.objects.create(
                    partner=payout.partner,
                    transaction_type='debit',
                    amount=payout.amount,
                    description=f'Payout processed - {remarks}'
                )
                
                messages.success(request, 'Payout completed and amount deducted from wallet!')
            else:
                messages.error(request, 'Payout must be approved first!')
                
        elif action == 'reject':
            payout.status = 'rejected'
            payout.processed_by = request.user
            payout.processed_at = timezone.now()
            payout.remarks = remarks
            payout.save()
            messages.warning(request, 'Payout rejected!')
    
    return redirect('admin_payouts')


# Partner ke views.py mein ye function add karo

# Partner Update Own Lead Stage (Limited Access)


   # Partner Update Own Lead (Sirf apni own leads update kar sakta hai)
# Partner Update Own Lead (Sirf apni own leads update kar sakta hai)
@login_required
def partner_update_lead(request, lead_id):
    if request.user.is_staff:
        return redirect('admin_dashboard')
    
    partner = get_object_or_404(Partner, user=request.user)
    lead = get_object_or_404(Lead, id=lead_id, partner=partner, lead_type='partner_own', assigned_to_admin=False)
    
    if request.method == 'POST':
        # ✅ Update customer details (AB SAB FIELDS EDITABLE HAIN)
        lead.customer_name = request.POST.get('customer_name')
        lead.customer_email = request.POST.get('customer_email')
        lead.customer_phone = request.POST.get('customer_phone')
        
        # ✅ Update lead details
        stage_id = request.POST.get('stage')
        lead.stage = LeadStage.objects.get(id=stage_id) if stage_id else None
        
        deal_amount = request.POST.get('deal_amount', 0)
        lead.deal_amount = float(deal_amount) if deal_amount else 0
        
        lead.save()
        
        # ✅ Handle notes (agar new note add kiya hai)
        notes_text = request.POST.get('notes', '').strip()
        if notes_text:
            LeadNote.objects.create(
                lead=lead,
                note=notes_text,
                created_by=request.user
            )
        
        messages.success(request, 'Lead updated successfully!')
        return redirect('partner_leads')
    
    return redirect('partner_leads')
# ✅ NAYA FUNCTION - Partner assigns lead to admin
@login_required
def partner_assign_lead_to_admin(request, lead_id):
    """Partner apni own lead ko admin ko assign karta hai"""
    if request.user.is_staff:
        return redirect('admin_dashboard')
    
    partner = get_object_or_404(Partner, user=request.user)
    lead = get_object_or_404(Lead, id=lead_id, partner=partner, lead_type='partner_own')
    
    # Check: Already assigned hai kya?
    if lead.assigned_to_admin:
        messages.warning(request, '⚠️ This lead is already assigned to admin!')
        return redirect('partner_leads')
    
    if request.method == 'POST':
        # Lead ko admin ko assign karo
        lead.assigned_to_admin = True
        lead.assigned_date = timezone.now()
        lead.save()
        
        messages.success(request, f'✅ Lead "{lead.customer_name}" assigned to admin successfully! You can now only view this lead.')
        return redirect('partner_leads')
    
    return redirect('partner_leads')



# Partner Delete Own Lead (Sirf unassigned leads delete kar sakta hai)
@login_required
def partner_delete_lead(request, lead_id):
    if request.user.is_staff:
        return redirect('admin_dashboard')
    
    partner = get_object_or_404(Partner, user=request.user)
    lead = get_object_or_404(Lead, id=lead_id, partner=partner, lead_type='partner_own')
    
    # Check: Agar lead assigned hai to delete nahi kar sakta
    if lead.assigned_to_admin:
        messages.error(request, '❌ Cannot delete! This lead is assigned to admin.')
        return redirect('partner_leads')
    
    if request.method == 'POST':
        lead_name = lead.customer_name
        lead.delete()
        messages.success(request, f'✅ Lead "{lead_name}" deleted successfully!')
        return redirect('partner_leads')
    
    return redirect('partner_leads')


# Admin Delete Lead
@login_required
def admin_delete_lead(request, lead_id):
    if not request.user.is_staff:
        return redirect('partner_login')
    
    lead = get_object_or_404(Lead, id=lead_id)
    
    if request.method == 'POST':
        lead_name = lead.customer_name
        lead.delete()
        messages.success(request, f'✅ Lead "{lead_name}" deleted successfully!')
        return redirect('admin_leads')
    
    return redirect('admin_leads')




from django.utils.text import slugify
from .models import Blog

# ============ ADMIN BLOG VIEWS ============

@login_required
def admin_blogs(request):
    if not request.user.is_staff:
        return redirect('partner_login')
    
    blogs = Blog.objects.all().order_by('order_number', '-created_at')
    
    context = {
        'blogs': blogs,
    }
    
    return render(request, 'referal_system/admin_blogs.html', context)


@login_required
def admin_add_blog(request):
    if not request.user.is_staff:
        return redirect('partner_login')
    
    if request.method == 'POST':
        title = request.POST.get('title')
        content = request.POST.get('content')
        order_number = request.POST.get('order_number', 0)
        status = request.POST.get('status', 'active')
        meta_description = request.POST.get('meta_description', '')
        featured_image = request.FILES.get('featured_image')
        
        # Auto-generate slug
        slug = slugify(title)
        
        # Check if slug exists
        if Blog.objects.filter(slug=slug).exists():
            slug = f"{slug}-{timezone.now().timestamp()}"
        
        Blog.objects.create(
            title=title,
            slug=slug,
            content=content,
            order_number=int(order_number),
            status=status,
            meta_description=meta_description,
            featured_image=featured_image,
            created_by=request.user
        )
        
        messages.success(request, '✅ Blog post created successfully!')
        return redirect('admin_blogs')
    
    return render(request, 'referal_system/admin_add_blog.html')


@login_required
def admin_edit_blog(request, blog_id):
    if not request.user.is_staff:
        return redirect('partner_login')
    
    blog = get_object_or_404(Blog, id=blog_id)
    
    if request.method == 'POST':
        blog.title = request.POST.get('title')
        blog.content = request.POST.get('content')
        blog.order_number = int(request.POST.get('order_number', 0))
        blog.status = request.POST.get('status', 'active')
        blog.meta_description = request.POST.get('meta_description', '')
        
        # Update slug if title changed
        new_slug = slugify(blog.title)
        if new_slug != blog.slug:
            if Blog.objects.filter(slug=new_slug).exclude(id=blog_id).exists():
                new_slug = f"{new_slug}-{timezone.now().timestamp()}"
            blog.slug = new_slug
        
        # Update image if provided
        if request.FILES.get('featured_image'):
            blog.featured_image = request.FILES.get('featured_image')
        
        blog.save()
        
        messages.success(request, '✅ Blog post updated successfully!')
        return redirect('admin_blogs')
    
    context = {
        'blog': blog,
    }
    
    return render(request, 'referal_system/admin_edit_blog.html', context)


@login_required
def admin_delete_blog(request, blog_id):
    if not request.user.is_staff:
        return redirect('partner_login')
    
    blog = get_object_or_404(Blog, id=blog_id)
    
    if request.method == 'POST':
        blog_title = blog.title
        blog.delete()
        messages.success(request, f'✅ Blog "{blog_title}" deleted successfully!')
        return redirect('admin_blogs')
    
    return redirect('admin_blogs')


@login_required
def admin_toggle_blog_status(request, blog_id):
    if not request.user.is_staff:
        return redirect('partner_login')
    
    blog = get_object_or_404(Blog, id=blog_id)
    blog.status = 'inactive' if blog.status == 'active' else 'active'
    blog.save()
    
    status_text = "activated" if blog.status == 'active' else "deactivated"
    messages.success(request, f'✅ Blog "{blog.title}" {status_text} successfully!')
    return redirect('admin_blogs')


# ============ PARTNER BLOG VIEWS ============

@login_required
def partner_blogs(request):
    if request.user.is_staff:
        return redirect('admin_dashboard')
    
    partner = get_object_or_404(Partner, user=request.user)
    
    # Partner sirf active blogs dekh sakta hai
    blogs = Blog.objects.filter(status='active').order_by('order_number', '-created_at')
    
    context = {
        'partner': partner,
        'blogs': blogs,
    }
    
    return render(request, 'referal_system/partner_blogs.html', context)


@login_required
def partner_blog_detail(request, blog_id):
    if request.user.is_staff:
        return redirect('admin_dashboard')
    
    partner = get_object_or_404(Partner, user=request.user)
    blog = get_object_or_404(Blog, id=blog_id, status='active')
    
    context = {
        'partner': partner,
        'blog': blog,
    }
    
    return render(request, 'referal_system/partner_blog_detail.html', context)



from .models import TeamMember, LeadNote

# ============ ADMIN TEAM MEMBER VIEWS ============

@login_required
def admin_team_members(request):
    if not request.user.is_staff:
        return redirect('partner_login')
    
    if request.method == 'POST':
        # Create new team member
        name = request.POST.get('name')
        email = request.POST.get('email')
        phone = request.POST.get('phone')
        username = request.POST.get('username')
        password = request.POST.get('password')
        role = request.POST.get('role')
        employee_id = request.POST.get('employee_id', '')
        department = request.POST.get('department', '')
        
        if User.objects.filter(username=username).exists():
            messages.error(request, 'Username already exists!')
        elif TeamMember.objects.filter(email=email).exists():
            messages.error(request, 'Email already exists!')
        else:
            user = User.objects.create_user(username=username, password=password, email=email)
            TeamMember.objects.create(
                user=user,
                name=name,
                email=email,
                phone=phone,
                role=role,
                employee_id=employee_id,
                department=department
            )
            messages.success(request, 'Team member added successfully!')
            return redirect('admin_team_members')
    
    team_members = TeamMember.objects.all().order_by('-created_at')
    
    context = {
        'team_members': team_members,
    }
    
    return render(request, 'referal_system/admin_team_members.html', context)


@login_required
def admin_edit_team_member(request, member_id):
    if not request.user.is_staff:
        return redirect('partner_login')
    
    member = get_object_or_404(TeamMember, id=member_id)
    
    if request.method == 'POST':
        member.name = request.POST.get('name')
        member.email = request.POST.get('email')
        member.phone = request.POST.get('phone')
        member.role = request.POST.get('role')
        member.employee_id = request.POST.get('employee_id', '')
        member.department = request.POST.get('department', '')
        member.status = request.POST.get('status')
        
        # Update username
        new_username = request.POST.get('username')
        if new_username != member.user.username:
            if User.objects.filter(username=new_username).exists():
                messages.error(request, 'Username already exists!')
                return redirect('admin_edit_team_member', member_id=member_id)
            member.user.username = new_username
            member.user.save()
        
        # Update password if provided
        new_password = request.POST.get('password')
        if new_password:
            member.user.set_password(new_password)
            member.user.save()
        
        member.save()
        messages.success(request, 'Team member updated successfully!')
        return redirect('admin_team_members')
    
    context = {
        'member': member,
    }
    
    return render(request, 'referal_system/admin_edit_team_member.html', context)


@login_required
def admin_delete_team_member(request, member_id):
    if not request.user.is_staff:
        return redirect('partner_login')
    
    member = get_object_or_404(TeamMember, id=member_id)
    
    if request.method == 'POST':
        member_name = member.name
        user = member.user
        member.delete()
        user.delete()
        messages.success(request, f'Team member "{member_name}" deleted successfully!')
        return redirect('admin_team_members')
    
    return redirect('admin_team_members')


@login_required
def admin_toggle_team_member_status(request, member_id):
    if not request.user.is_staff:
        return redirect('partner_login')
    
    member = get_object_or_404(TeamMember, id=member_id)
    member.status = 'inactive' if member.status == 'active' else 'active'
    member.save()
    
    status_text = "activated" if member.status == 'active' else "deactivated"
    messages.success(request, f'Team member "{member.name}" {status_text} successfully!')
    return redirect('admin_team_members')


# ============ LEAD ASSIGNMENT & NOTES ============

@login_required
def admin_assign_team_member(request, lead_id):
    """Assign team member to lead"""
    if not request.user.is_staff:
        return redirect('partner_login')
    
    lead = get_object_or_404(Lead, id=lead_id)
    
    if request.method == 'POST':
        team_member_id = request.POST.get('team_member')
        
        if team_member_id:
            team_member = TeamMember.objects.get(id=team_member_id)
            lead.assigned_team_member = team_member
            lead.save()
            messages.success(request, f'Lead assigned to {team_member.name} successfully!')
        else:
            lead.assigned_team_member = None
            lead.save()
            messages.success(request, 'Team member unassigned from lead!')
        
        return redirect('admin_leads')
    
    return redirect('admin_leads')

@login_required
def admin_add_lead_note(request, lead_id):
    """Admin adds follow-up note with next follow-up date"""
    if not request.user.is_staff:
        return redirect('partner_login')
    
    lead = get_object_or_404(Lead, id=lead_id)
    
    if request.method == 'POST':
        note_text = request.POST.get('note')
        follow_up_date = request.POST.get('follow_up_date')
        
        # Create note
        LeadNote.objects.create(
            lead=lead,
            note=note_text,
            
        )
        
        # ✅ UPDATE next_follow_up date if provided
        if follow_up_date:
            lead.next_follow_up = follow_up_date
            lead.save()
            messages.success(request, f'✅ Note added! Next follow-up set for {follow_up_date}')
        else:
            messages.success(request, '✅ Note added successfully!')
        
        return redirect('admin_lead_detail', lead_id=lead.id)
    
    return redirect('admin_lead_detail', lead_id=lead.id)




@login_required
def admin_lead_detail(request, lead_id):
    """View lead details with notes and assignments"""
    if not request.user.is_staff:
        return redirect('partner_login')
    
    lead = get_object_or_404(Lead, id=lead_id)
    notes = lead.notes.all().order_by('-created_at')
    team_members = TeamMember.objects.filter(status='active')
    stages = LeadStage.objects.all()
    
    context = {
        'lead': lead,
        'notes': notes,
        'team_members': team_members,
        'stages': stages,
    }
    
    return render(request, 'referal_system/admin_lead_detail.html', context)



from django.db.models import Q, Count
from datetime import date, timedelta

# ============ TEAM MEMBER VIEWS ============

@login_required
def team_dashboard(request):
    """Team Member Dashboard"""
    if request.user.is_staff:
        return redirect('admin_dashboard')
    
    if not hasattr(request.user, 'teammember'):
        messages.error(request, 'You are not registered as a team member')
        return redirect('partner_login')
    
    team_member = request.user.teammember
    
    # Stats
    total_assigned_leads = Lead.objects.filter(assigned_team_member=team_member).count()
    
    # Today's follow-ups
    today = date.today()
    today_followups = LeadNote.objects.filter(
        team_member=team_member,
        follow_up_date=today,
        follow_up_completed=False
    ).count()
    
    # Pending follow-ups (overdue)
    pending_followups = LeadNote.objects.filter(
        team_member=team_member,
        follow_up_date__lt=today,
        follow_up_completed=False
    ).count()
    
    # Upcoming follow-ups (next 7 days)
    next_week = today + timedelta(days=7)
    upcoming_followups = LeadNote.objects.filter(
        team_member=team_member,
        follow_up_date__range=[today, next_week],
        follow_up_completed=False
    ).count()
    
    # Recent assigned leads
    recent_leads = Lead.objects.filter(assigned_team_member=team_member).order_by('-created_at')[:10]
    
    context = {
        'team_member': team_member,
        'total_assigned_leads': total_assigned_leads,
        'today_followups': today_followups,
        'pending_followups': pending_followups,
        'upcoming_followups': upcoming_followups,
        'recent_leads': recent_leads,
    }
    
    return render(request, 'referal_system/team_dashboard.html', context)


@login_required
def team_my_leads(request):
    """Team Member - My Assigned Leads"""
    if not hasattr(request.user, 'teammember'):
        return redirect('partner_login')
    
    team_member = request.user.teammember
    
    # Get filter parameter
    filter_type = request.GET.get('filter', 'all')
    
    leads = Lead.objects.filter(assigned_team_member=team_member)
    
    today = date.today()
    
    # Apply filters
    if filter_type == 'today':
        # Leads with today's follow-up
        lead_ids = LeadNote.objects.filter(
            team_member=team_member,
            follow_up_date=today,
            follow_up_completed=False
        ).values_list('lead_id', flat=True)
        leads = leads.filter(id__in=lead_ids)
    
    elif filter_type == 'pending':
        # Leads with overdue follow-ups
        lead_ids = LeadNote.objects.filter(
            team_member=team_member,
            follow_up_date__lt=today,
            follow_up_completed=False
        ).values_list('lead_id', flat=True)
        leads = leads.filter(id__in=lead_ids)
    
    elif filter_type == 'upcoming':
        # Leads with upcoming follow-ups (next 7 days)
        next_week = today + timedelta(days=7)
        lead_ids = LeadNote.objects.filter(
            team_member=team_member,
            follow_up_date__range=[today, next_week],
            follow_up_completed=False
        ).values_list('lead_id', flat=True)
        leads = leads.filter(id__in=lead_ids)
    
    leads = leads.order_by('-created_at')
    stages = LeadStage.objects.all()
    
    context = {
        'team_member': team_member,
        'leads': leads,
        'stages': stages,
        'filter_type': filter_type,
    }
    
    return render(request, 'referal_system/team_my_leads.html', context)

@login_required
def team_lead_detail(request, lead_id):
    """Team Member - Lead Detail with Notes"""
    if not hasattr(request.user, 'teammember'):
        return redirect('partner_login')
    
    team_member = request.user.teammember
    lead = get_object_or_404(Lead, id=lead_id, assigned_team_member=team_member)
    
    notes = lead.notes.filter(team_member=team_member).order_by('-created_at')
    stages = LeadStage.objects.all()
    
    # ✅ Prepare upcoming follow-ups in view
    upcoming_followups = lead.notes.filter(
        team_member=team_member,
        follow_up_date__isnull=False,
        follow_up_completed=False
    ).order_by('follow_up_date')[:5]
    
    context = {
        'team_member': team_member,
        'lead': lead,
        'notes': notes,
        'stages': stages,
        'upcoming_followups': upcoming_followups,  # ✅ Add this
        'today': date.today(),  # ✅ Add this for date input
    }
    
    return render(request, 'referal_system/team_lead_detail.html', context)

@login_required
def team_update_lead(request, lead_id):
    """Team Member - Update Lead Status"""
    if not hasattr(request.user, 'teammember'):
        return redirect('partner_login')
    
    team_member = request.user.teammember
    lead = get_object_or_404(Lead, id=lead_id, assigned_team_member=team_member)
    
    if request.method == 'POST':
        stage_id = request.POST.get('stage')
        deal_amount = request.POST.get('deal_amount')
        
        if stage_id:
            lead.stage = LeadStage.objects.get(id=stage_id)
        
        if deal_amount:
            lead.deal_amount = float(deal_amount)
        
        lead.save()
        
        messages.success(request, '✅ Lead updated successfully!')
        return redirect('team_lead_detail', lead_id=lead_id)
    
    return redirect('team_my_leads')


@login_required
def team_add_note(request, lead_id):
    """Team Member - Add Follow-up Note"""
    if not hasattr(request.user, 'teammember'):
        return redirect('partner_login')
    
    team_member = request.user.teammember
    lead = get_object_or_404(Lead, id=lead_id, assigned_team_member=team_member)
    
    if request.method == 'POST':
        note_text = request.POST.get('note')
        follow_up_date = request.POST.get('follow_up_date')
        
        if note_text:
            LeadNote.objects.create(
                lead=lead,
                team_member=team_member,
                note=note_text,
                follow_up_date=follow_up_date if follow_up_date else None
            )
            messages.success(request, '✅ Follow-up note added successfully!')
        
        return redirect('team_lead_detail', lead_id=lead_id)
    
    return redirect('team_my_leads')


@login_required
def team_mark_followup_complete(request, note_id):
    """Team Member - Mark Follow-up as Complete"""
    if not hasattr(request.user, 'teammember'):
        return redirect('partner_login')
    
    team_member = request.user.teammember
    note = get_object_or_404(LeadNote, id=note_id, team_member=team_member)
    
    note.follow_up_completed = True
    note.save()
    
    messages.success(request, '✅ Follow-up marked as complete!')
    return redirect('team_lead_detail', lead_id=note.lead.id)


@login_required
def team_followup_calendar(request):
    """Team Member - Follow-up Calendar View"""
    if not hasattr(request.user, 'teammember'):
        return redirect('partner_login')
    
    team_member = request.user.teammember
    
    # Get all follow-ups for this team member
    followups = LeadNote.objects.filter(
        team_member=team_member,
        follow_up_date__isnull=False
    ).order_by('follow_up_date')
    
    # Group by date
    from collections import defaultdict
    calendar_data = defaultdict(list)
    
    for followup in followups:
        calendar_data[followup.follow_up_date].append(followup)
    
    context = {
        'team_member': team_member,
        'calendar_data': dict(calendar_data),
    }
    
    return render(request, 'referal_system/team_followup_calendar.html', context)



import pandas as pd
from django.contrib import messages
from django.shortcuts import render, redirect
from datetime import datetime

import pandas as pd
from django.contrib import messages
from django.shortcuts import render, redirect
from datetime import datetime
from django.db.models import Count

import pandas as pd
from django.contrib import messages
from django.shortcuts import render, redirect
from datetime import datetime
from django.db.models import Count

def admin_bulk_upload_leads(request):
    """Bulk upload leads via Excel file with Round Robin Assignment"""
    
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        lead_type = request.POST.get('lead_type', 'admin')
        
        if not excel_file:
            messages.error(request, 'Please select an Excel file')
            return redirect('admin_bulk_upload_leads')
        
        try:
            # Read Excel file
            df = pd.read_excel(excel_file)
            
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Validate required fields
                    if pd.isna(row.get('customer_name')) or pd.isna(row.get('customer_email')):
                        errors.append(f"Row {index+2}: Missing required fields")
                        error_count += 1
                        continue
                    
                    partner = None
                    
                    if lead_type == 'admin':
                        # ✅ ADMIN LEAD - Khud ke liye, NO PARTNER
                        partner = None
                    
                    elif lead_type == 'partner':
                        # ✅ PARTNER LEAD - Partners ko assign
                        if not pd.isna(row.get('partner_email')):
                            # Sheet mein partner email diya hai
                            try:
                                partner = Partner.objects.get(
                                    email=str(row['partner_email']).strip(),
                                    is_verified=True,
                                    is_active=True
                                )
                            except Partner.DoesNotExist:
                                errors.append(f"Row {index+2}: Partner not found or not active - {row.get('partner_email')}")
                                error_count += 1
                                continue
                        else:
                            # ✅ Sheet mein partner nahi diya - Round Robin Logic
                            partner = get_next_partner_round_robin()
                            
                            if not partner:
                                errors.append(f"Row {index+2}: No active partners available for assignment")
                                error_count += 1
                                continue
                    
                    # Get stage if provided
                    stage = None
                    if not pd.isna(row.get('stage_name')):
                        try:
                            stage = LeadStage.objects.get(name=row['stage_name'])
                        except LeadStage.DoesNotExist:
                            pass
                    
                    # Create Lead
                    Lead.objects.create(
                        partner=partner,
                        customer_name=str(row['customer_name']).strip(),
                        customer_email=str(row['customer_email']).strip(),
                        customer_phone=str(row.get('customer_phone', '')).strip(),
                        lead_type=lead_type,
                        stage=stage,
                        deal_amount=float(row.get('deal_amount', 0)),
                        commission_percent=float(row.get('commission_percent', 0)) if lead_type == 'partner' else 0,
                        additional_notes=str(row.get('notes', '')),
                        assigned_to_admin=True if lead_type == 'admin' else False,
                        assigned_date=datetime.now()
                    )
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {index+2}: {str(e)}")
                    error_count += 1
            
            # Success message
            if success_count > 0:
                messages.success(request, f'✅ Successfully uploaded {success_count} leads!')
            
            if error_count > 0:
                error_msg = f'❌ {error_count} leads failed. Errors: ' + ' | '.join(errors[:10])
                messages.warning(request, error_msg)
            
            return redirect('admin_leads')
            
        except Exception as e:
            messages.error(request, f'Error reading file: {str(e)}')
            return redirect('admin_bulk_upload_leads')
    
    # GET request - show form
    partners = Partner.objects.filter(is_verified=True, is_active=True)
    stages = LeadStage.objects.all()
    
    context = {
        'partners': partners,
        'stages': stages,
    }
    return render(request, 'referal_system/admin_bulk_upload.html', context)


def get_next_partner_round_robin():
    """
    Round Robin Logic - Sabse kam leads wale partner ko assign karo
    """
    partners = Partner.objects.filter(
        is_verified=True,
        is_active=True
    ).annotate(
        lead_count=Count('lead')
    ).order_by('lead_count', 'created_at')
    
    if partners.exists():
        return partners.first()
    
    return None


from django.http import HttpResponse
import openpyxl
from openpyxl.styles import Font, PatternFill

def download_sample_excel(request):
    """Generate and download sample Excel file"""
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # ===== Sheet 1: Admin Leads =====
    ws1 = wb.active
    ws1.title = "Admin Leads Sample"
    
    headers = [
        'customer_name', 'customer_email', 'customer_phone', 
        'stage_name', 'deal_amount', 'notes'
    ]
    
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    # Admin leads sample data (NO partner_email column)
    admin_data = [
        ['Rahul Kumar', 'rahul@example.com', '9876543210', 'New', '50000', 'Admin own lead'],
        ['Priya Sharma', 'priya@example.com', '9876543211', 'Contacted', '75000', 'Direct lead'],
    ]
    
    for row_num, row_data in enumerate(admin_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws1.cell(row=row_num, column=col_num, value=value)
    
    # ===== Sheet 2: Partner Leads =====
    ws2 = wb.create_sheet("Partner Leads Sample")
    
    partner_headers = [
        'customer_name', 'customer_email', 'customer_phone', 
        'partner_email', 'stage_name', 'deal_amount', 
        'commission_percent', 'notes'
    ]
    
    for col_num, header in enumerate(partner_headers, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    # Partner leads sample data
    partner_data = [
        ['Amit Singh', 'amit@example.com', '9876543212', 'partner1@example.com', 'New', '60000', '10', 'Specific partner'],
        ['Neha Gupta', 'neha@example.com', '9876543213', '', 'Contacted', '80000', '10', 'Auto Round Robin - Empty email'],
        ['Vikram Shah', 'vikram@example.com', '9876543214', '', 'New', '70000', '10', 'Auto assigned'],
    ]
    
    for row_num, row_data in enumerate(partner_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws2.cell(row=row_num, column=col_num, value=value)
    
    # Add instruction row in Partner sheet
    ws2.insert_rows(2)
    instruction_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    instruction_font = Font(italic=True, color="E67E22", bold=True)
    
    instruction_cell = ws2.cell(row=2, column=1)
    instruction_cell.value = "⚠️ Leave partner_email EMPTY for automatic Round Robin assignment | Fill partner_email to assign specific partner"
    instruction_cell.fill = instruction_fill
    instruction_cell.font = instruction_font
    ws2.merge_cells('A2:H2')
    
    # Adjust column widths for both sheets
    for ws in [ws1, ws2]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=lead_upload_samples.xlsx'
    wb.save(response)
    
    return response







@login_required
def partner_bulk_upload_leads(request):
    """Partner bulk upload - Own leads ya Referral leads"""
    
    if request.user.is_staff:
        return redirect('admin_dashboard')
    
    partner = get_object_or_404(Partner, user=request.user)
    
    if request.method == 'POST':
        excel_file = request.FILES.get('excel_file')
        lead_ownership = request.POST.get('lead_ownership', 'own')
        
        if not excel_file:
            messages.error(request, 'Please select an Excel file')
            return redirect('partner_bulk_upload_leads')
        
        try:
            # Read Excel file
            df = pd.read_excel(excel_file)
            
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Validate required fields
                    if pd.isna(row.get('customer_name')) or pd.isna(row.get('customer_email')):
                        errors.append(f"Row {index+2}: Missing required fields")
                        error_count += 1
                        continue
                    
                    # Get stage if provided
                    stage = None
                    if not pd.isna(row.get('stage_name')):
                        try:
                            stage = LeadStage.objects.get(name=row['stage_name'])
                        except LeadStage.DoesNotExist:
                            pass
                    
                    # Create Lead based on ownership
                    if lead_ownership == 'own':
                        # Own Lead
                        lead = Lead.objects.create(
                            lead_type='partner_own',
                            partner=partner,
                            customer_name=str(row['customer_name']).strip(),
                            customer_email=str(row['customer_email']).strip(),
                            customer_phone=str(row.get('customer_phone', '')).strip(),
                            stage=stage,
                            deal_amount=float(row.get('deal_amount', 0)),
                            commission_percent=0,
                            assigned_to_admin=False
                        )
                    else:
                        # Referral Lead
                        lead = Lead.objects.create(
                            lead_type='partner_referral',
                            partner=partner,
                            customer_name=str(row['customer_name']).strip(),
                            customer_email=str(row['customer_email']).strip(),
                            customer_phone=str(row.get('customer_phone', '')).strip(),
                            stage=stage,
                            deal_amount=float(row.get('deal_amount', 0)),
                            commission_percent=float(row.get('commission_percent', 0)),
                            assigned_to_admin=True,
                            assigned_date=datetime.now()
                        )
                    
                    # Add notes if present
                    notes_text = str(row.get('notes', '')).strip()
                    if notes_text and notes_text != 'nan':
                        LeadNote.objects.create(
                            lead=lead,
                            note=notes_text,
                            created_by=request.user
                        )
                    
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f"Row {index+2}: {str(e)}")
                    error_count += 1
            
            # Success message
            if success_count > 0:
                if lead_ownership == 'own':
                    messages.success(request, f'✅ Successfully uploaded {success_count} own leads!')
                else:
                    messages.success(request, f'✅ Successfully uploaded {success_count} referral leads to admin!')
            
            if error_count > 0:
                error_msg = f'❌ {error_count} leads failed. Errors: ' + ' | '.join(errors[:10])
                messages.warning(request, error_msg)
            
            return redirect('partner_leads')
            
        except Exception as e:
            messages.error(request, f'Error reading file: {str(e)}')
            return redirect('partner_bulk_upload_leads')
    
    # GET request - show form
    stages = LeadStage.objects.all()
    
    context = {
        'partner': partner,
        'stages': stages,
    }
    return render(request, 'referal_system/partner_bulk_upload.html', context)

@login_required
def partner_download_sample_excel(request):
    """Partner sample Excel file download"""
    
    if request.user.is_staff:
        return redirect('admin_dashboard')
    
    # Create workbook
    wb = openpyxl.Workbook()
    
    # ===== Sheet 1: Own Leads =====
    ws1 = wb.active
    ws1.title = "Own Leads Sample"
    
    headers = [
        'customer_name', 'customer_email', 'customer_phone', 
        'stage_name', 'deal_amount', 'notes'
    ]
    
    header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    
    for col_num, header in enumerate(headers, 1):
        cell = ws1.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    # Own leads sample data
    own_data = [
        ['Rahul Kumar', 'rahul@example.com', '9876543210', 'New', '50000', 'My own lead'],
        ['Priya Sharma', 'priya@example.com', '9876543211', 'Contacted', '75000', 'Direct contact'],
    ]
    
    for row_num, row_data in enumerate(own_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws1.cell(row=row_num, column=col_num, value=value)
    
    # ===== Sheet 2: Referral Leads =====
    ws2 = wb.create_sheet("Referral Leads Sample")
    
    referral_headers = [
        'customer_name', 'customer_email', 'customer_phone', 
        'stage_name', 'deal_amount', 'commission_percent', 'notes'
    ]
    
    for col_num, header in enumerate(referral_headers, 1):
        cell = ws2.cell(row=1, column=col_num)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
    
    # Referral leads sample data
    referral_data = [
        ['Amit Singh', 'amit@example.com', '9876543212', 'New', '60000', '10', 'Referral to admin'],
        ['Neha Gupta', 'neha@example.com', '9876543213', 'Contacted', '80000', '12', 'High value lead'],
    ]
    
    for row_num, row_data in enumerate(referral_data, 2):
        for col_num, value in enumerate(row_data, 1):
            ws2.cell(row=row_num, column=col_num, value=value)
    
    # Add instruction row in Referral sheet
    ws2.insert_rows(2)
    instruction_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    instruction_font = Font(italic=True, color="E67E22", bold=True)
    
    instruction_cell = ws2.cell(row=2, column=1)
    instruction_cell.value = "💰 Admin will manage these leads and you will get commission when closed!"
    instruction_cell.fill = instruction_fill
    instruction_cell.font = instruction_font
    ws2.merge_cells('A2:G2')
    
    # Adjust column widths for both sheets
    for ws in [ws1, ws2]:
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = max_length + 2
    
    # Response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=partner_lead_upload_samples.xlsx'
    wb.save(response)
    
    return response